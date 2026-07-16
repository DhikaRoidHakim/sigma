import csv
import io
import qrcode
from pathlib import Path
from openpyxl import Workbook, load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas as pdfcanvas
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

PRIMARY = colors.HexColor("#01567A")
BORDER = colors.HexColor("#E5E7EB")
LOGO_PATH = Path(__file__).parent / "logo.png"

ASSET_COLUMNS = [
    ("kode_aset", "Kode Aset"),
    ("nama_aset", "Nama Aset"),
    ("jenis_inventaris", "Jenis Inventaris"),
    ("golongan", "Golongan"),
    ("tanggal_pembelian", "Tanggal Pembelian"),
    ("nilai_pembelian", "Nilai Pembelian"),
    ("status", "Status"),
    ("current_office_name", "Kantor"),
    ("current_room_name", "Ruangan"),
]

HISTORY_COLUMNS = [
    ("moved_at", "Tanggal"),
    ("from_office_name", "Dari Kantor"),
    ("from_room_name", "Dari Ruangan"),
    ("to_office_name", "Ke Kantor"),
    ("to_room_name", "Ke Ruangan"),
    ("moved_by", "Dipindahkan Oleh"),
    ("notes", "Catatan"),
]

REPAIR_COLUMNS = [
    ("tanggal_perbaikan", "Tanggal"),
    ("deskripsi_kerusakan", "Deskripsi Kerusakan"),
    ("tindakan_perbaikan", "Tindakan Perbaikan"),
    ("biaya", "Biaya"),
    ("teknisi", "Teknisi/Vendor"),
    ("status", "Status"),
    ("created_by", "Dicatat Oleh"),
]


def _fmt(value):
    if value is None:
        return ""
    return str(value)


def qr_png_bytes(data: str) -> bytes:
    qr = qrcode.QRCode(box_size=10, border=2, error_correction=qrcode.constants.ERROR_CORRECT_M)
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill_color="#01567A", back_color="white")
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()


def _truncate_to_width(c, text: str, font: str, size: float, max_width: float) -> str:
    if c.stringWidth(text, font, size) <= max_width:
        return text
    while text and c.stringWidth(text + "…", font, size) > max_width:
        text = text[:-1]
    return text + "…"


def _draw_label(c, x: float, y: float, w: float, h: float, asset: dict, qr_data: str):
    c.setStrokeColor(BORDER)
    c.setLineWidth(0.6)
    c.roundRect(x, y, w, h, 2.5 * mm, stroke=1, fill=0)
    pad = 3.5 * mm
    qr_size = h - 2 * pad
    qr_reader = ImageReader(io.BytesIO(qr_png_bytes(qr_data)))
    c.drawImage(qr_reader, x + pad, y + pad, qr_size, qr_size)

    tx = x + pad + qr_size + 3.5 * mm
    text_w = x + w - pad - tx
    top = y + h - pad

    logo_size = 5.5 * mm
    if LOGO_PATH.exists():
        c.drawImage(ImageReader(str(LOGO_PATH)), tx, top - logo_size, logo_size, logo_size, mask="auto")
        brand_x = tx + logo_size + 1.8 * mm
    else:
        brand_x = tx
    c.setFillColor(PRIMARY)
    c.setFont("Helvetica-Bold", 9)
    c.drawString(brand_x, top - logo_size + 1.4 * mm, "SIGMA")

    c.setFillColor(colors.black)
    c.setFont("Courier-Bold", 11)
    c.drawString(tx, top - logo_size - 6 * mm, _truncate_to_width(c, asset.get("kode_aset", ""), "Courier-Bold", 11, text_w))

    c.setFillColor(colors.HexColor("#1F2937"))
    c.setFont("Helvetica", 7.5)
    nama = asset.get("nama_aset", "")
    line1 = _truncate_to_width(c, nama, "Helvetica", 7.5, text_w)
    c.drawString(tx, top - logo_size - 10.5 * mm, line1)

    c.setFillColor(colors.HexColor("#6B7280"))
    c.setFont("Helvetica", 6.5)
    lokasi = " / ".join(filter(None, [asset.get("current_office_name"), asset.get("current_room_name")])) or "Belum ditempatkan"
    c.drawString(tx, top - logo_size - 14.5 * mm, _truncate_to_width(c, lokasi, "Helvetica", 6.5, text_w))


def single_label_pdf(asset: dict, qr_data: str) -> bytes:
    w, h = 100 * mm, 40 * mm
    buf = io.BytesIO()
    c = pdfcanvas.Canvas(buf, pagesize=(w, h))
    _draw_label(c, 1.5 * mm, 1.5 * mm, w - 3 * mm, h - 3 * mm, asset, qr_data)
    c.showPage()
    c.save()
    return buf.getvalue()


def bulk_labels_pdf(assets: list, qr_data_fn) -> bytes:
    page_w, page_h = A4
    label_w, label_h = 90 * mm, 36 * mm
    cols, rows = 2, 7
    margin_x = (page_w - cols * label_w) / (cols + 1)
    margin_y = (page_h - rows * label_h) / (rows + 1)
    buf = io.BytesIO()
    c = pdfcanvas.Canvas(buf, pagesize=A4)
    for i, asset in enumerate(assets):
        pos = i % (cols * rows)
        if i > 0 and pos == 0:
            c.showPage()
        col = pos % cols
        row = pos // cols
        x = margin_x + col * (label_w + margin_x)
        y = page_h - margin_y - label_h - row * (label_h + margin_y)
        _draw_label(c, x, y, label_w, label_h, asset, qr_data_fn(asset))
    c.showPage()
    c.save()
    return buf.getvalue()


def rows_to_csv(rows: list, columns: list) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([label for _, label in columns])
    for row in rows:
        writer.writerow([_fmt(row.get(key)) for key, _ in columns])
    return buf.getvalue().encode("utf-8-sig")


def rows_to_xlsx(rows: list, columns: list, sheet_title: str = "Data") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.append([label for _, label in columns])
    for row in rows:
        ws.append([row.get(key) if row.get(key) is not None else "" for key, _ in columns])
    for i, (_, label) in enumerate(columns, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = max(16, len(label) + 4)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def rows_to_pdf(rows: list, columns: list, title: str, subtitle: str = "") -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=15 * mm, bottomMargin=15 * mm,
                            leftMargin=12 * mm, rightMargin=12 * mm)
    styles = getSampleStyleSheet()
    cell_style = ParagraphStyle("cell", parent=styles["Normal"], fontSize=8, leading=10)
    header_style = ParagraphStyle("header", parent=styles["Normal"], fontSize=8, leading=10,
                                  textColor=colors.white, fontName="Helvetica-Bold")
    elements = [
        Paragraph(f"SIGMA — {title}", ParagraphStyle("title", parent=styles["Title"], fontSize=16, textColor=PRIMARY)),
    ]
    if subtitle:
        elements.append(Paragraph(subtitle, ParagraphStyle("sub", parent=styles["Normal"], fontSize=9, textColor=colors.HexColor("#6B7280"))))
    elements.append(Spacer(1, 6 * mm))

    data = [[Paragraph(label, header_style) for _, label in columns]]
    for row in rows:
        data.append([Paragraph(_fmt(row.get(key)), cell_style) for key, _ in columns])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), PRIMARY),
        ("GRID", (0, 0), (-1, -1), 0.5, BORDER),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(table)
    doc.build(elements)
    return buf.getvalue()


IMPORT_FIELDS = ["kode_aset", "nama_aset", "jenis_inventaris", "golongan",
                 "tanggal_pembelian", "nilai_pembelian", "status", "nama_kantor", "nama_ruangan"]

HEADER_ALIASES = {label.lower(): key for key, label in ASSET_COLUMNS}
HEADER_ALIASES.update({key: key for key in IMPORT_FIELDS})
HEADER_ALIASES.update({"kantor": "nama_kantor", "ruangan": "nama_ruangan",
                       "nama_kantor": "nama_kantor", "nama_ruangan": "nama_ruangan"})


def _normalize_headers(headers: list) -> list:
    return [HEADER_ALIASES.get(str(h or "").strip().lower(), str(h or "").strip().lower()) for h in headers]


def parse_upload(filename: str, content: bytes) -> list:
    name = (filename or "").lower()
    if name.endswith(".xlsx"):
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            headers = _normalize_headers(list(next(rows_iter)))
        except StopIteration:
            return []
        rows = []
        for raw in rows_iter:
            if raw is None or all(v is None or str(v).strip() == "" for v in raw):
                continue
            rows.append({headers[i]: raw[i] for i in range(min(len(headers), len(raw)))})
        return rows
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    all_rows = [r for r in reader if any(str(c).strip() for c in r)]
    if not all_rows:
        return []
    headers = _normalize_headers(all_rows[0])
    return [{headers[i]: row[i] for i in range(min(len(headers), len(row)))} for row in all_rows[1:]]
